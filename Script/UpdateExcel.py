import requests

from CSFloat import get_skins_bought, get_skins_sold
from DealsDB import get_sell_range_by_id

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from copy import copy
from datetime import date

# Red fill
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Grey fill
grey_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
# Green fill
green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

# Fetches the current USD → EUR rate from an online API
def get_usd_to_eur_rate():
    url = "https://open.er-api.com/v6/latest/USD"

    response = requests.get(url, timeout=10)
    data = response.json()

    if data.get("result") != "success":
        raise Exception("Error fetching rate")

    return data["rates"]["EUR"]



# Converts a price from USD to EUR
# Parameters:
#   - usd_value: price in USD
#   - rate: current USD → EUR rate
def convert_usd_to_eur(usd_value, rate):
    return round(usd_value * rate, 2)



# Builds rows of new bought skins to insert into the Excel file that aren't on it yet
# Parameters:
#   - last_id_in_excel: ID of the last bought skin already registered in the Excel file
def extract_info_from_skins_bought(last_id_in_excel, rate):
    
    # 200 newest bought skins
    skins_bought = get_skins_bought()
    
    rows_for_excel = []
    for skin in skins_bought:
        skin_id = skin["contract_id"]
        bought_at, _ = skin["created_at"].split("T") # e.g.: 2026-02-10
        # Can stop extracting details from skins bought because from now on all those are already in the Excel file
        if skin_id == last_id_in_excel:
            break

        sell_for = get_sell_range_by_id(skin_id) or "N/A"

        # Build an entire row to add in the Excel with the information about a bought skin
        row = {
            "name": skin["contract"]["item"]["market_hash_name"],
            "id": skin_id,
            "float": f"{skin["contract"]["item"]["float_value"]:.12f}",
            "bought_at": bought_at,
            "bought_for$": skin["contract"]["price"]/100,
            "bought_for€": convert_usd_to_eur(skin["contract"]["price"]/100, rate),
            "sell_for$(when_was_bought)": sell_for
        }

        # Insert the row in the first position of the list
        rows_for_excel.insert(0, row)
    return rows_for_excel



# Returns the amount of days it took to sell a skin based on the date where it was bought and the date where it was sold
# Parameters:
#   - bought_at_date: date where the skin was bought (E.g.: 2025-10-03)
#   - sold_at_date: date where the skin was sold (E.g.: 2025-10-25)
def days_to_sell(bought_at_date, sold_at_date):
    new_bought_at_date = date.fromisoformat(bought_at_date)
    new_sold_at_date = date.fromisoformat(sold_at_date)
    return abs((new_sold_at_date - new_bought_at_date).days) - 8 # '-8' because after the skin is bought, it remains unsellable for an entire week



# Returns an entire Excel row with the information about a bought skin completed with the information about its sale
# Parameters:
#   - trades_by_contract_skin_key: dictionary where the key is the concatenation of the skin name and its float value, and the value is the trade information of a sold skin
#   - row: dictionary containing the information about a bought skin (excel row)
#   - rate: current USD → EUR rate
def complete_row(trades_by_contract_skin_key, row, rate):
    key = f"{row['name']}|{row['float']}"
    if trades_by_contract_skin_key.get(key):
        sold_at, _ = trades_by_contract_skin_key[key]["created_at"].split("T")
        row["days_to_sell"] = days_to_sell(row["bought_at"], sold_at)
        row["sold_for$(taxes_included)"] = trades_by_contract_skin_key[key]["contract"]["price"]/100
        row["sold_for€(taxes_included)"] = convert_usd_to_eur(row["sold_for$(taxes_included)"], rate)
        row["profit$(post_taxes)"] = row["sold_for$(taxes_included)"]*0.98 - row["bought_for$"]
        row["profit€(post_taxes)"] = convert_usd_to_eur(row["profit$(post_taxes)"], rate)
        row["percentage_of_profit"] = (row["profit$(post_taxes)"] / row["bought_for$"])
    return row



# Completes the information about incompleted rows in the Excel worksheet
# Parameters:
#   - header_map: dictionary where the key is the column name and the value is the column number (e.g.: {"name": 1, "id": 2, ...})
#   - worksheet: Excel worksheet object where the rows to complete are
#   - trades_by_contract_skin_key: dictionary where the key is the concatenation of the skin name and its float value, and the value is the trade information of a sold skin
#   - rate: current USD → EUR rate
def update_incompleted_rows_from_excel(header_map, worksheet, trades_by_contract_skin_key, rate):
    # Get the column number of each column name
    name_column = header_map["name"]
    id_column = header_map["id"]
    float_column = header_map["float"]
    bought_at_column = header_map["bought_at"]
    days_to_sell_column = header_map["days_to_sell"]
    bought_for_USD_column = header_map["bought_for$"]
    sell_for_column = header_map["sell_for$(when_was_bought)"]

    rows_updated = 0
    for row in range(2, worksheet.max_row + 1):
        # Get the ID and the days_to_sell value of the current row (skin)
        id_value = worksheet.cell(row=row, column=id_column).value
        days_to_sell_value = worksheet.cell(row=row, column=days_to_sell_column).value

        # If row has an ID but no days_to_sell yet, the row needs to be updated
        if id_value and days_to_sell_value is None:

            # Get the rest of values of the row
            name_value = worksheet.cell(row=row, column=name_column).value
            float_value = worksheet.cell(row=row, column=float_column).value
            bought_at_value = worksheet.cell(row=row, column=bought_at_column).value
            bought_for_USD_value = worksheet.cell(row=row, column=bought_for_USD_column).value
            sell_for_value = worksheet.cell(row=row, column=sell_for_column).value

            row_data = {
                "name": name_value,
                "id": id_value,
                "float": float_value,
                "bought_at": bought_at_value,
                "bought_for$": bought_for_USD_value,
                "bought_for€": convert_usd_to_eur(bought_for_USD_value, rate),
                "sell_for$(when_was_bought)": sell_for_value
            }
            row_data = complete_row(trades_by_contract_skin_key, row_data, rate)
            if "days_to_sell" in row_data:
                rows_updated += 1
            # Update the row in the Excel with the completed information
            for header, value in row_data.items():
                col = header_map[header]
                worksheet.cell(row=row, column=col, value=value)
    return rows_updated



# Completes the information about new bought skins to insert into the Excel file that aren't on it yet 
# with the information about their sale (if they were sold) 
# and updates the previous incompleted rows in the Excel file with the information about their sale (if they were sold)
# Parameters:
#   - rows_to_append: rows of new bought skins to insert into the Excel file that aren't on it yet
#   - header_map: dictionary where the key is the column name and the value is the column number (e.g.: {"name": 1, "id": 2, ...})
#   - worksheet: Excel worksheet object where the rows to complete are
def extract_info_from_skins_sold(rows_to_append, header_map, worksheet, rate):
    skins_sold = get_skins_sold()
    trades_by_contract_skin_key = {
        f"{trade['contract']['item']['market_hash_name']}|{trade['contract']['item']['float_value']:.12f}": trade
        for trade in skins_sold if trade.get('contract', {}).get('item', {}).get('float_value') is not None
    }
    # Complete the information about the new bought skins to insert into the Excel file that aren't on it yet 
    # with the information about their sale (if they were sold)
    if rows_to_append:
        for i, row in enumerate(rows_to_append):
            key = f"{row['name']}|{row['float']}"
            trade = trades_by_contract_skin_key.get(key)
            if trade:
                rows_to_append[i] = complete_row(trades_by_contract_skin_key, row, rate)
    
    # Complete the information about incompleted rows in the Excel worksheet with the information about their sale (if they were sold)
    rows_updated = update_incompleted_rows_from_excel(header_map, worksheet, trades_by_contract_skin_key, rate)

    return rows_to_append, rows_updated



# Main function that updates the Excel file with the information about new bought/sold skins
# Parameters:
#   - file_path: path to the Excel file to update
async def update_excel(file_path: str):
    # Load existing Excel file
    workbook = load_workbook(file_path, keep_vba=True)
    try:
        worksheet = workbook["Trades"]

        # Map column names -> column number
        header_map = {}
        for column in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=column).value
            if header:
                header_map[header] = column

        id_column = header_map["id"]
        last_row = None # row where there is data written

        # Find the last row with data
        for row in range(worksheet.max_row, 1, -1):  # start from bottom
            cell_value = worksheet.cell(row=row, column=id_column).value
            if cell_value is not None and cell_value != "":
                last_row = row # It will fall here if there are already rows filled in the Excel
                break

        # Load the last skin ID in Excel
        if last_row is not None:
            last_id_in_excel = str(worksheet.cell(row=last_row, column=id_column).value)
        else:
            last_id_in_excel = None

        rate = get_usd_to_eur_rate()
        # Get new rows to insert
        rows = extract_info_from_skins_bought(last_id_in_excel, rate)
        rows, rows_updated = extract_info_from_skins_sold(rows, header_map, worksheet, rate)
        if not rows:
            workbook.save(file_path)
            return {"rows_added": 0, "rows_updated": rows_updated}

        # Detect the table
        table = list(worksheet._tables.values())[0]

        start_cell, end_cell = table.ref.split(":")
        table_last_row = int(end_cell[1:])

        # Insert new row before the bottom row
        if last_row is None:  # Here it means that the Excel hasn't got skins so there is an entire empty row
            rows_to_insert = len(rows) - 1
            if rows_to_insert > 0:
                worksheet.insert_rows(table_last_row, amount=rows_to_insert)
        else:
            worksheet.insert_rows(table_last_row, amount=len(rows))

        # Last line before the bottom line (where the 'Total Profit' is)
        template_row = table_last_row - 1

        if last_row is None:  # Here it means that the Excel hasn't got skins so there is an entire empty row
            range_to_use = len(rows) - 1
        else:
            range_to_use = len(rows)
        for i in range(range_to_use):
            new_row = table_last_row + i
            for col in range(1, worksheet.max_column + 1):
                source = worksheet.cell(row=template_row, column=col)
                target = worksheet.cell(row=new_row, column=col)
                if source.has_style:
                    target.font = copy(source.font)
                    target.border = copy(source.border)
                    target.fill = copy(source.fill)
                    target.number_format = copy(source.number_format)
                    target.protection = copy(source.protection)
                    target.alignment = copy(source.alignment)

        # Write new rows
        if last_row is None: # Here it means that the Excel hasn't got skins so there is an entire empty row
            write_row = table_last_row - 1
        else:
            write_row = table_last_row
        for row_data in rows:
            for header, value in row_data.items():
                col = header_map[header]
                worksheet.cell(row=write_row, column=col, value=value)
            write_row += 1

        # Expand the table
        if last_row is None: # Here it means that the Excel hasn't got skins so there is an entire empty row
            new_last_row = table_last_row + len(rows) - 1
        else:
            new_last_row = table_last_row + len(rows)
        table.ref = f"{start_cell}:{end_cell[0]}{new_last_row}"

        # Columns to apply rules
        columns = ["K", "L"]

        # Rules for painting cells based in a formula
        rules = [
            ("{col}2<0", red_fill),
            ("AND({col}2>=0,{col}2<=0.009)", grey_fill),
            ("{col}2>0.009", green_fill)
        ]

        # Remove previous rules in the column and add the new one
        worksheet.conditional_formatting._cf_rules.clear()

        for col in columns:
            for formula_template, fill in rules:
                formula = formula_template.format(col=col)  # e.g., "L2>0.009"
                rule = FormulaRule(formula=[formula], fill=fill)
                worksheet.conditional_formatting.add(f"{col}2:{col}{new_last_row-1}", rule)

        workbook.save(file_path)
        return {"rows_added": len(rows), "rows_updated": rows_updated}
    finally:
        workbook.close()